import openpyxl
from openpyxl import Workbook
import os

# File to store the material data in Excel format
EXCEL_FILE = 'minecraft_build_castle_materials.xlsx'

# Create a new Excel file (if it doesn't exist)
def create_excel_file():
    wb = Workbook()
    ws = wb.active
    ws.title = "Materials"
    
    # Create headers
    ws.append(["Material", "Total Required", "Gathered", "Remaining"])

    wb.save(EXCEL_FILE)
    return

# Load the Excel file (create it if not exists)
def load_excel():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        return wb
    except FileNotFoundError:
        create_excel_file()
        return openpyxl.load_workbook(EXCEL_FILE)
    return

def sort_materials():
    wb = load_excel()
    ws = wb["Materials"]

    # Collect all rows of data, excluding the header
    data = list(ws.iter_rows(min_row=2, max_col=4, values_only=True))
    
    # Sort data alphabetically by the first column (Material)
    data.sort(key=lambda row: row[0] if row[0] else "")

    # Clear existing rows but keep the header
    ws.delete_rows(2, ws.max_row - 1)

    # Write sorted rows back to the worksheet
    for row_index, row_data in enumerate(data, start=2):
        for col_index, value in enumerate(row_data, start=1):
            ws.cell(row=row_index, column=col_index, value=value)
    
    wb.save(EXCEL_FILE)
    #print("Materials sorted alphabetically.")
    return

# Update materials in the Excel file
def update_material(material_name, total_required, gathered_quantity):
    wb = load_excel()
    ws = wb["Materials"]

    # Check if material already exists in the file
    material_found = False
    for row in ws.iter_rows(min_row=2, max_col=4):
        if row[0].value == material_name:
            row[1].value = total_required  # Update total required
            row[2].value = gathered_quantity  # Update gathered
            row[3].value = total_required - gathered_quantity  # Update remaining
            material_found = True
            break

    # If material not found, add a new row
    if not material_found:
        remaining = total_required - gathered_quantity
        ws.append([material_name, total_required, gathered_quantity, remaining])

    wb.save(EXCEL_FILE)
    print(f"{material_name} updated: {gathered_quantity}/{total_required} (Remaining: {total_required - gathered_quantity})")
    
    # Sort materials after updating
    sort_materials()
    return

# Display all materials in the Excel file
def display_materials():
    wb = load_excel()
    ws = wb["Materials"]

    print("Current Materials:")
    for row in ws.iter_rows(min_row=2, values_only=True):
        print(f"Material: {row[0]}, Total Required: {row[1]}, Gathered: {row[2]}, Remaining: {row[3]}")
    return

def main():
    # "Item Name", Numbers of Items Needed, Numbers of Items
    
    # Page 1/3
    update_material("Calcite", 2129, 98)
    update_material('Cherry Log', 673, 304)
    update_material("Bamboo", 154, 192)
    update_material("Stone", 182, 1152)
    update_material("Stone Brick", 270, 372)
    update_material("Stone Brick Slab", 0, 16)
    update_material("Cobbled Deepslate", 0, 2112)
    update_material("Polished Deepslate Slab", 612, 0)
    update_material("Polished Deepslate Wall", 332, 0)
    update_material("Polished Deepslate Stairs", 138, 0)
    update_material("Polished Andesite", 1807, 394)
    update_material("Cherry Sapling", 66, 118)
    update_material("Andesite", 495, 80)
    update_material("Polished Andesite Slab", 362, 0)
    update_material("Cherry Trapdoor", 359, 0)
    update_material("Cherry Fence", 354, 0)
    update_material("Stone Stairs", 342, 0)
    update_material("Cherry Fence Gate", 314, 0)
    update_material("Crimson Stairs", 303, 0)
    update_material("Crimson Planks", 300, 0)
    update_material("Candle", 169, 0)
    update_material("Cherry Leaves", 169, 192)
    update_material("Polished Diorite", 164, 0)
    update_material("Cherry Slab", 141, 0)
    update_material("Cracked Stone Bricks", 140, 0)
    update_material("Polished Deepslate Stairs", 138, 0)
    update_material("Lantern", 121, 0)
    update_material("Stone Brick Wall", 117, 0)
    update_material("Mossy Stone Bricks", 101, 0)
    update_material("Andesite Stairs", 98, 3)
    update_material("Grass Block", 97, 0)
    update_material("Red Carpet", 91, 96)
    update_material("Flower Pot", 88, 0)
    update_material("Chain", 56, 0)
    update_material("Lightning Rod", 52, 0)
    update_material("Chest", 43, 43)
    
    # Page 2/3
    
    update_material("Andesite Slab", 39, 0)
    update_material("Decorated Pot", 39, 0)
    update_material("Red Wool", 32, 128)
    update_material("Dark Oak Trapdoor", 30, 0)
    update_material("Diorite Wall", 29, 0)
    update_material("Dark Oak Stairs", 28, 0)
    update_material("Red Candle", 28, 0)
    update_material("Bookshelf", 27, 0)
    update_material("Cherry Hanging Sign", 26, 0)
    update_material("Torch", 24, 0)
    update_material("Polished Andesite Stairs", 23, 0)    
    update_material("Water Bucket", 22, 0)
    update_material("Barrel", 20, 0)
    update_material("Yellow Candle", 20, 0)
    update_material("Andesite Wall", 19, 0)
    update_material("Crimson Fence", 18, 0)
    update_material("Stripped Dark Oak Log", 17, 0)
    update_material("Cherry Door", 12, 0)
    update_material("Pink Carpet", 12, 63)
    update_material("Yellow Banner", 10, 0)
    update_material("Yellow Carpet", 10, 96)
    update_material("Cherry Planks", 9, 0)
    update_material("Campfire", 8, 0)
    update_material("Polished Deepslate", 6, 0)
    update_material("Scaffolding", 6, 0)
    update_material("Stripped Cherry Log", 6, 0)
    update_material("Shroomlight", 5, 24)
    update_material("Cauldron", 4, 0)
    update_material("Dandelion", 4, 14)
    update_material("Lectern", 4, 0)
    update_material("Sea Pickle", 4, 0)
    update_material("Azure Bluet", 3, 0)
    update_material("Cactus", 3, 0)
    update_material("Pink Bed", 3, 0)
    update_material("Poppy", 3, 106)
    update_material("Brewing Stand", 2, 0)
    
    # Page 3/3
    update_material("Cake", 2, 0)
    update_material("Crimson Fungus", 2, 0)
    update_material("Dark Oak Planks", 2, 0)
    update_material("Fern", 2, 0)
    update_material("Flowering Azalea", 2, 0)
    update_material("Jukebox", 2, 0)
    update_material("Podzol", 2, 0)
    update_material("Skeleton Skull", 2, 0)
    update_material("String", 2, 0)
    update_material("Blast Furnace", 1, 0)
    update_material("Crafting Table", 1, 1)
    update_material("Enchantment Table", 1, 0)
    update_material("End Rod", 1, 0)
    update_material("Grindstone", 1, 0)
    update_material("Large Amethyst Bud", 1, 0)
    update_material("Oxeye Daisy", 1, 0)
    update_material("Smithing Table", 1, 0)
    update_material("Stonecutter", 1, 0)
    
    # extra?
    
    update_material("Diorite", 0, 45)
    update_material("Pink Wool", 0, 128)
    update_material("Yellow Wool", 0, 128)
    update_material("Crimson Stem", 0, 194)
    update_material("Crimson Roots", 0, 1)
    update_material("Weeping Vines", 0, 6)
    update_material("Red Mushroom", 0, 1)
    update_material("Stick", 0, 72)
    update_material("Nether Wart Block", 0, 108)
    update_material("Bone Block", 0, 17)
    update_material("Bone Meal", 0, 11)
    update_material("Leather", 0, 3)
    
    display_materials()
    return

main()
print(os.path.abspath(EXCEL_FILE))
